import openpyxl as px


# 設定ファイルの読み込み
def read_setting(file_path):
  workbook = px.load_workbook(file_path)
  sheet =workbook.active
  return sheet["C5"].value,sheet["D5"].value

# ファイルからチケットを検索
def read_excel(file_path, number):
  workbook = px.load_workbook(file_path)
  sheet = workbook.worksheets[0]
  
  for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
    if row[0] == number:
      print("見つけた")

  print("Finish!")

# 設定ファイルのパスを設定
file_path = '抽出設定.xlsx'
path,number =  read_setting(file_path)
read_excel(path, number)
